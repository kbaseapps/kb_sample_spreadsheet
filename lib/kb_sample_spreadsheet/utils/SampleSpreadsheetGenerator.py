import os
import time
import jsonschema
from pprint import pprint

from kb_sample_spreadsheet.utils import schema, column
from installed_clients.KBaseReportClient import KBaseReport

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle, PatternFill, Font

_SCHEMAS=schema.load_schemas()
_COLUMNS=column.load_columns()

class SampleSpreadsheetGenerator(object):

  def __init__(self, scratch_dir, callback_url, ctx):
    self.scratch_dir = scratch_dir
    self.callback_url = callback_url
    self.provenance = ctx.provenance()
    pass

  def generate(self, params):
    validated_params = self.validate_params(params, 'sampleSpreadsheetParams')
    spreashsheet_info=self.generate_spreadsheet(validated_params, 'IGSN')
    report_info=self.create_report(validated_params, spreashsheet_info)
    return report_info

  def validate_params(self, params, schema_name):
    _params = _params=params.copy()
    _schema = _SCHEMAS[schema_name]
    try:
      jsonschema.validate(instance=_params, schema=_schema)
    except ValidationError as err:
      raise ValueError('Parameter validation error: ' + err.message)
    return _params

  def generate_spreadsheet(self, validated_params, template_name):
    spreadsheet_info={}

    output_dir=os.path.join(self.scratch_dir, 'sample_spreadsheet_' + str(int(time.time() * 10000)))
    output_filename=validated_params['output_name'] + '.xlsx'
    output_filepath=os.path.join(output_dir, output_filename)
    os.makedirs(output_dir)
   
    _column=_COLUMNS[template_name]

    common_font=Font(name='Arial', size=9)
    highlight_font=Font(name='Arial', size=9, bold=True)
    header_fill=PatternFill("solid", fgColor="D1E5FE")

    wb = Workbook()
    ws = wb.active
    ws.title = "samples"
    ws.page_setup.fitToWidth = 1

    ws['A1']='Object Type:'
    ws['A1'].font=highlight_font
    ws['B1']=validated_params['object_type']
    ws['B1'].font=common_font
    ws['C1']='User Code:'
    ws['C1'].font=highlight_font
    ws['D1']=validated_params['user_code']
    ws['D1'].font=common_font

    for c in range(1, len(_column)+1):
      _col=_column[c-1]
      cell=ws.cell(row=2, column=c, value=_col['header'])
      cell.fill=header_fill
      cell.font=highlight_font
      if 'type' in _col and _col['type']:
        cl=get_column_letter(c)
        dv=DataValidation(type=_col['type'], allow_blank=True)
        dv.error ='Your entry is not ' + _col['type']
        dv.errorTitle = 'Invalid Entry'
        ws.add_data_validation(dv)
        dv.add(cl + '3:' + cl + '1048576')

    wb.save(filename = output_filepath)
    spreadsheet_info['output_dir']=output_dir
    spreadsheet_info['output_file']=[{
      'path': output_filepath,
      'name': output_filename,
      'label': output_filename,
      'description': 'Sample spreadsheet'
      }]
    return spreadsheet_info
  
  def create_report(self, validated_params, spreadsheet_info):
    kbr = KBaseReport(self.callback_url)
    report_params={
      'file_links': spreadsheet_info['output_file'],
      'workspace_name': validated_params['workspace_name']
    }
    report_info=kbr.create_extended_report(report_params)
    return {'report_name': report_info['name'], 'report_ref': report_info['ref']}
